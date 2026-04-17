#!/usr/bin/env python3
"""
Usage:
    python generate_report.py monthly_data.json daily_data.json [output.xlsx]

JSON schemas
    monthly : [{"department":str, "subdivision":str, "usage":num, "amount":num}, ...]
    daily   : [{"date":str, "department":str, "subdivision":str,
                "usage":num, "amount":num}, ...]
"""

import io, json, sys
from collections import OrderedDict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ═══════════════════════════════════════════════════════════════════════════════

ACCENT  = "B6D0E2"   # user-specified accent colour
WHITE   = "FFFFFF"
BLACK   = "000000"

ACCENT_FILL = PatternFill("solid", fgColor=ACCENT)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)

# Fonts – all black text
TITLE_FONT  = Font(name="Arial", bold=True,  color=BLACK, size=13)
HDR_FONT    = Font(name="Arial", bold=True,  color=BLACK, size=10)
DATA_FONT   = Font(name="Arial", bold=False, color=BLACK, size=10)
TOT_FONT    = Font(name="Arial", bold=True,  color=BLACK, size=10)

# Alignments
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center")
RGT  = Alignment(horizontal="right",  vertical="center")

# Borders – thin outer only; used only on individual data cells
_thin  = Side(style="thin",   color="BFBFBF")
_none  = Side(style=None)
CELL_BDR  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NO_BDR    = Border(left=_none, right=_none, top=_none, bottom=_none)

# Number formats
NUM_FMT = "#,##0"
AMT_FMT = "#,##0.00"

# ─── Chart palette ────────────────────────────────────────────────────────────
C_BAR  = "#B6D0E2"
C_LINE = "#FF0000"
C_BG   = "#FFFFFF"
C_GRID = "#E0E0E0"
C_ACC  = "#" + ACCENT

plt.rcParams.update({
    "font.family":     "DejaVu Sans",
    "axes.spines.top": False,
    "axes.grid":       True,
    "grid.color":      C_GRID,
    "grid.linestyle":  "--",
    "grid.linewidth":  0.5,
})

# ═══════════════════════════════════════════════════════════════════════════════
#  BORDER UTILITY  – apply clean outer border to a merged range
# ═══════════════════════════════════════════════════════════════════════════════

def _outer_border(ws, r1, c1, r2, c2, side=_thin):
    """Apply border only to the outer edges of a cell range – no inner lines."""
    top    = Border(top=side)
    bot    = Border(bottom=side)
    left   = Border(left=side)
    right  = Border(right=side)
    tl     = Border(top=side, left=side)
    tr_    = Border(top=side, right=side)
    bl     = Border(bottom=side, left=side)
    br_    = Border(bottom=side, right=side)

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            t = side if r == r1 else _none
            b = side if r == r2 else _none
            l = side if c == c1 else _none
            ri = side if c == c2 else _none
            cell.border = Border(top=t, bottom=b, left=l, right=ri)


def _fill_range(ws, r1, c1, r2, c2, fill, font=None, align=None):
    """Apply fill/font/alignment to every cell in a range."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            if font:  cell.font      = font
            if align: cell.alignment = align


def _merge(ws, r1, c1, r2, c2, value, fill, font, align=CTR):
    """Merge a range, write value in top-left, apply fill across range, clean outer border."""
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    _fill_range(ws, r1, c1, r2, c2, fill, font, align)
    ws.cell(r1, c1).value     = value
    ws.cell(r1, c1).alignment = align
    _outer_border(ws, r1, c1, r2, c2)


def _cell(ws, row, col, value=None, fill=WHITE_FILL, font=DATA_FONT,
          align=LEFT, border=CELL_BDR, num_fmt=None):
    c = ws.cell(row, col)
    if value  is not None: c.value         = value
    c.fill      = fill
    c.font      = font
    c.alignment = align
    c.border    = border
    if num_fmt:            c.number_format = num_fmt
    return c


def _col_widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════════
#  CHART GENERATORS  (return PNG bytes)
# ═══════════════════════════════════════════════════════════════════════════════

def _chart_table(ax, labels, usage, amount):
    """Render a small value table beneath the x-axis."""
    tbl = ax.table(
        cellText  = [[f"{v:,.0f}" for v in usage],
                     [f"{v:,.0f}" for v in amount]],
        rowLabels = ["Usage", "Amount"],
        colLabels = labels,
        cellLoc   = "center",
        loc       = "bottom",
        bbox      = [0, -0.40, 1, 0.33],
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    for (r, c), cell in tbl.get_celld().items():
        cell.set_linewidth(0.4)
        cell.set_edgecolor("#CCCCCC")
        if r == 0:                          # col-label row
            cell.set_facecolor(C_ACC)
            cell.set_text_props(color="black", fontweight="bold")
        elif r == 1:                        # Usage row
            cell.set_facecolor(C_BAR)
            cell.set_text_props(color="black")
        else:                               # Amount row
            cell.set_facecolor("#FFFFFF")
            cell.set_text_props(color="black")
        if c == -1:                         # row-label col
            cell.set_facecolor(C_ACC)
            cell.set_text_props(color="black", fontweight="bold")


def _render_combo(labels, usage, amount, title, fig_w=13, fig_h=7.2, bottom_pad=0.40):
    fig, ax1 = plt.subplots(figsize=(fig_w, fig_h))
    fig.patch.set_facecolor(C_BG)
    plt.subplots_adjust(bottom=bottom_pad, left=0.09, right=0.91, top=0.91)

    x  = np.arange(len(labels))
    bw = 0.55

    # ── Usage bars ────────────────────────────────────────────────────────────
    bars = ax1.bar(x, usage, width=bw, color=C_BAR, zorder=3,
                   linewidth=0.8, edgecolor="#7AAECB", label="Usage")
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=20, ha="right", fontsize=9, color="black")
    ax1.set_ylabel("Usage", color="black", fontsize=10, fontweight="bold")
    ax1.tick_params(axis="y", labelcolor="black")
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax1.set_facecolor(C_BG)
    ax1.spines["left"].set_color("#AAAAAA")
    ax1.spines["bottom"].set_color("#AAAAAA")

    # Usage value labels inside bars (bottom)
    for bar, val in zip(bars, usage):
        if bar.get_height() > 0:
            ax1.text(bar.get_x() + bar.get_width() / 2,
                     bar.get_height() * 0.04,
                     f"{val:,.0f}",
                     ha="center", va="bottom",
                     fontsize=7.5, color="black", fontweight="bold", zorder=5)

    # ── Amount line – secondary axis ──────────────────────────────────────────
    ax2 = ax1.twinx()
    ax2.plot(x, amount, color=C_LINE, linewidth=2.4,
             marker="o", markersize=6,
             markerfacecolor=C_LINE, markeredgecolor="#AA0000",
             label="Amount", zorder=4)
    ax2.set_ylabel("Amount", color="black", fontsize=10, fontweight="bold")
    ax2.tick_params(axis="y", labelcolor="black")
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax2.spines["right"].set_visible(True)
    ax2.spines["right"].set_color("#AAAAAA")
    ax2.spines["top"].set_visible(False)

    # Amount value labels on top of line
    for xi, val in zip(x, amount):
        ax2.text(xi, val * 1.018, f"{val:,.0f}",
                 ha="center", va="bottom",
                 fontsize=7.5, color=C_LINE, fontweight="bold", zorder=5)

    ax1.set_title(title, fontsize=12, fontweight="bold", color="black", pad=12)

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="upper left", fontsize=9, framealpha=0.9)

    _chart_table(ax1, labels, usage, amount)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=C_BG)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def chart_png_monthly(monthly_data):
    return _render_combo(
        [r["subdivision"] for r in monthly_data],
        [r["usage"]       for r in monthly_data],
        [r["amount"]      for r in monthly_data],
        "Monthly Usage & Amount by Subdivision", fig_w=13, fig_h=7.2)


def chart_png_summary_total(monthly_data):
    return _render_combo(
        [r["subdivision"] for r in monthly_data],
        [r["usage"]       for r in monthly_data],
        [r["amount"]      for r in monthly_data],
        "Total Monthly Usage & Amount Overview", fig_w=14, fig_h=7.5)


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 – MONTHLY
# ═══════════════════════════════════════════════════════════════════════════════

def build_sheet_monthly(ws, monthly_data):
    ws.title = "Monthly Data"
    NCOLS = 4

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _merge(ws, 1, 1, 1, NCOLS, "Monthly Data", ACCENT_FILL, TITLE_FONT)
    ws.row_dimensions[1].height = 32

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for col, lbl in enumerate(["Department", "Sub Division", "Usage", "Amount"], 1):
        _merge(ws, 2, col, 2, col, lbl, ACCENT_FILL, HDR_FONT)
    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    start_row = 3
    for i, rec in enumerate(monthly_data):
        r = start_row + i
        _cell(ws, r, 1, rec["department"],  align=LEFT)
        _cell(ws, r, 2, rec["subdivision"], align=LEFT)
        _cell(ws, r, 3, rec["usage"],       align=RGT, num_fmt=NUM_FMT)
        _cell(ws, r, 4, rec["amount"],      align=RGT, num_fmt=AMT_FMT)
        ws.row_dimensions[r].height = 18

    last_data = start_row + len(monthly_data) - 1
    total_row = last_data + 1

    # ── Total row ─────────────────────────────────────────────────────────────
    _merge(ws, total_row, 1, total_row, 2, "Total", ACCENT_FILL, TOT_FONT)
    _cell(ws, total_row, 3, f"=SUM(C{start_row}:C{last_data})",
          fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=NUM_FMT)
    _cell(ws, total_row, 4, f"=SUM(D{start_row}:D{last_data})",
          fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=AMT_FMT)
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = "A3"
    _col_widths(ws, {"A": 18, "B": 18, "C": 14, "D": 16})

    # ── Chart below total ─────────────────────────────────────────────────────
    png = chart_png_monthly(monthly_data)
    buf = io.BytesIO(png)
    img = XLImage(buf)
    img.width, img.height = 830, 510
    ws.add_image(img, f"A{total_row + 2}")
    print("  ✓ Monthly chart embedded")


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 – DAILY
# ═══════════════════════════════════════════════════════════════════════════════

def build_sheet_daily(ws, daily_data):
    ws.title = "Daily Data"

    subdivisions = list(OrderedDict.fromkeys(r["subdivision"] for r in daily_data))
    n_subs    = len(subdivisions)
    sub_idx   = {s: i for i, s in enumerate(subdivisions)}
    tot_u_col = 2 + n_subs * 2
    tot_a_col = tot_u_col + 1
    NCOLS     = tot_a_col

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _merge(ws, 1, 1, 1, NCOLS, "Daily Data", ACCENT_FILL, TITLE_FONT)
    ws.row_dimensions[1].height = 32

    # ── Row 2: Date | Subdivision names | Total ───────────────────────────────
    _merge(ws, 2, 1, 3, 1, "Date", ACCENT_FILL, HDR_FONT)   # spans rows 2-3
    for i, sub in enumerate(subdivisions):
        c = 2 + i * 2
        _merge(ws, 2, c, 2, c + 1, sub, ACCENT_FILL, HDR_FONT)
    _merge(ws, 2, tot_u_col, 2, tot_a_col, "Total", ACCENT_FILL, HDR_FONT)
    ws.row_dimensions[2].height = 22

    # ── Row 3: Usage / Amount sub-headers ─────────────────────────────────────
    for i in range(n_subs):
        c = 2 + i * 2
        _merge(ws, 3, c,     3, c,     "Usage",  ACCENT_FILL, HDR_FONT)
        _merge(ws, 3, c + 1, 3, c + 1, "Amount", ACCENT_FILL, HDR_FONT)
    _merge(ws, 3, tot_u_col, 3, tot_u_col, "Usage",  ACCENT_FILL, HDR_FONT)
    _merge(ws, 3, tot_a_col, 3, tot_a_col, "Amount", ACCENT_FILL, HDR_FONT)
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    by_date = OrderedDict()
    for rec in daily_data:
        by_date.setdefault(rec["date"], {})[rec["subdivision"]] = rec

    data_start = 4
    for r_idx, (date, subs_data) in enumerate(by_date.items()):
        row = data_start + r_idx
        _cell(ws, row, 1, date, align=CTR, font=HDR_FONT)
        usage_refs, amount_refs = [], []
        for sub in subdivisions:
            c = 2 + sub_idx[sub] * 2
            u = subs_data[sub]["usage"]  if sub in subs_data else 0
            a = subs_data[sub]["amount"] if sub in subs_data else 0
            _cell(ws, row, c,     u, align=RGT, num_fmt=NUM_FMT)
            _cell(ws, row, c + 1, a, align=RGT, num_fmt=AMT_FMT)
            usage_refs.append(f"{get_column_letter(c)}{row}")
            amount_refs.append(f"{get_column_letter(c+1)}{row}")
        _cell(ws, row, tot_u_col, "=" + "+".join(usage_refs),
              fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=NUM_FMT)
        _cell(ws, row, tot_a_col, "=" + "+".join(amount_refs),
              fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=AMT_FMT)
        ws.row_dimensions[row].height = 18

    last_data = data_start + len(by_date) - 1
    total_row = last_data + 1

    # ── Total row ─────────────────────────────────────────────────────────────
    _merge(ws, total_row, 1, total_row, 1, "Total", ACCENT_FILL, TOT_FONT)
    for i in range(n_subs):
        c  = 2 + i * 2
        uc = get_column_letter(c)
        ac = get_column_letter(c + 1)
        _cell(ws, total_row, c,     f"=SUM({uc}{data_start}:{uc}{last_data})",
              fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=NUM_FMT)
        _cell(ws, total_row, c + 1, f"=SUM({ac}{data_start}:{ac}{last_data})",
              fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=AMT_FMT)
    tuc = get_column_letter(tot_u_col)
    tac = get_column_letter(tot_a_col)
    _cell(ws, total_row, tot_u_col,
          f"=SUM({tuc}{data_start}:{tuc}{last_data})",
          fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=NUM_FMT)
    _cell(ws, total_row, tot_a_col,
          f"=SUM({tac}{data_start}:{tac}{last_data})",
          fill=ACCENT_FILL, font=TOT_FONT, align=RGT, num_fmt=AMT_FMT)
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 14
    for i in range(n_subs):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width     = 12
        ws.column_dimensions[get_column_letter(2 + i * 2 + 1)].width = 14
    ws.column_dimensions[get_column_letter(tot_u_col)].width = 13
    ws.column_dimensions[get_column_letter(tot_a_col)].width = 14
    print("  ✓ Daily sheet built")


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 – SUMMARY CHARTS
# ═══════════════════════════════════════════════════════════════════════════════

def build_sheet_summary(ws, monthly_data, daily_data):
    ws.title = "Summary Charts"
    NCOLS = 16

    # ── Title ─────────────────────────────────────────────────────────────────
    _merge(ws, 1, 1, 1, NCOLS, "Summary – Usage & Amount Overview",
           ACCENT_FILL, TITLE_FONT)
    ws.row_dimensions[1].height = 32

    # ── Total combo chart ─────────────────────────────────────────────────────
    png_total = chart_png_summary_total(monthly_data)
    buf = io.BytesIO(png_total)
    img = XLImage(buf)
    img.width, img.height = 980, 540
    ws.add_image(img, "A2")
    print("  ✓ Total chart embedded")

    # ── Section 2 title ───────────────────────────────────────────────────────
    section2_row = 32
    _merge(ws, section2_row, 1, section2_row, NCOLS,
           "Daily Breakdown by Subdivision", ACCENT_FILL,
           Font(name="Arial", bold=True, color=BLACK, size=12))
    ws.row_dimensions[section2_row].height = 26

    # ── Per-subdivision daily charts ──────────────────────────────────────────
    by_sub = OrderedDict()
    for r in daily_data:
        s = r["subdivision"]
        if s not in by_sub:
            by_sub[s] = {"dates": [], "usage": [], "amount": []}
        by_sub[s]["dates"].append(r["date"])
        by_sub[s]["usage"].append(r["usage"])
        by_sub[s]["amount"].append(r["amount"])

    CHART_ROWS = 22
    COLS       = ["A", "J"]

    for i, sub in enumerate(by_sub):
        d = by_sub[sub]
        png = _render_combo(
            [dt[-5:] for dt in d["dates"]],
            d["usage"], d["amount"],
            f"{sub} – Daily Usage & Amount",
            fig_w=10, fig_h=6.0, bottom_pad=0.42
        )
        buf = io.BytesIO(png)
        img = XLImage(buf)
        img.width, img.height = 680, 420
        anchor = f"{COLS[i % 2]}{section2_row + 1 + (i // 2) * CHART_ROWS}"
        ws.add_image(img, anchor)
        print(f"  ✓ {sub} chart at {anchor}")

    _col_widths(ws, {"A": 14})


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report(monthly_json, daily_json, output="report.xlsx"):
    print("Loading data …")
    with open(monthly_json) as f: monthly_data = json.load(f)
    with open(daily_json)   as f: daily_data   = json.load(f)

    wb = Workbook()
    print("Building Sheet 1 – Monthly Data …")
    build_sheet_monthly(wb.active, monthly_data)
    print("Building Sheet 2 – Daily Data …")
    build_sheet_daily(wb.create_sheet(), daily_data)
    print("Building Sheet 3 – Summary Charts …")
    build_sheet_summary(wb.create_sheet(), monthly_data, daily_data)
    wb.save(output)
    print(f"\n✓ Report saved → {output}")


if __name__ == "__main__":
    monthly_path = sys.argv[1] if len(sys.argv) > 1 else "monthly_data.json"
    daily_path   = sys.argv[2] if len(sys.argv) > 2 else "daily_data.json"
    out_path     = sys.argv[3] if len(sys.argv) > 3 else "report.xlsx"
    generate_report(monthly_path, daily_path, out_path)